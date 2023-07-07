from netmiko import ConnectHandler
from paramiko.ssh_exception import SSHException
from netmiko.ssh_exception import NetMikoTimeoutException
from netmiko import CNTL_SHIFT_6
from getpass import getpass
import os
import openpyxl
import re
from datetime import datetime
import sys
from time import time
from time import time
import logging

os.chdir('/Users/wenjiama/OneDrive - Procter and Gamble/work/netdev')
#os.chdir('C:\\Users\\ma.w.8\\OneDrive - Procter and Gamble\\work\\netdev')
print(os.getcwd())


def Login(host, type, username, password, port=22):
    device = {
        'host': host,
        'device_type': type,
        'username': username,
        'password': password
    }
    try:
        connection = ConnectHandler(**device)
        return connection
    except (EOFError, SSHException, NetMikoTimeoutException):
        # Set the filename of the output .txt file to the IP Address of this device
        print(host + 'cannot be connected')
        return

if __name__ == '__main__':
    device_workbook = openpyxl.load_workbook('./device_workbook.xlsx')
    device_list = device_workbook['device_list']
    row_count = device_list.max_row
    print(row_count)
    column_count = device_list.max_column
    print(column_count)
    for i in range(2, row_count+1):
        host = device_list['A'+str(i)].value
        type ='cisco_ios'
        username = 'ma.w.8'
        password = '9P7us9GS'
        ssh = Login(host, type, username, password)
        if ssh != None:
            start_1=time()
            devicename = re.search(r'(\S+)#', ssh.find_prompt()).group(1)
            print('login '+devicename+' successfully')
            device_folder = devicename+'_'+datetime.now().strftime('%Y-%m-%d %H')
            if os.path.exists('./command capture/output-3/' + device_folder) == True:
                print(device_folder+' folder exist')
            else:
                os.mkdir('./command capture/output-3/'+device_folder)
            #os.mkdir('./command capture/output-3/'+device_folder)
            #os.mkdir('./command capture\\output\\'+device_folder)
            command_list = './command_list.txt'
            with open(command_list, 'r') as command_list:
                commands = command_list.readlines()
            for i in commands:
                   command = re.search(r'(.+)\n', i).group(1)
                   print(command)
                   show_output = './command capture/output-3/'+device_folder+'/'+command+'.txt'
                   #show_output = './command capture\\output\\'+device_folder+'\\'+command+'.txt'
#                   logging.basicConfig(filename='test.log', level=logging.DEBUG)
#                   logger = logging.getLogger("netmiko")
                   read = ssh.read_channel()
                   try:
                       ssh.send_command('show ip routes')
                   except:
                       print(read)
#                   with open(show_output, 'w+') as show_output:
#                       try:
#                           show_output.write(ssh.send_command(i,delay_factor=2))
#                       except:
                   ssh.disconnect()
            start_2=time()
            print('run time : %.2f' % (start_2 - start_1))
        else:
            continue
